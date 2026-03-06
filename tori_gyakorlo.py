import random
import re
import tkinter as tk
from tkinter import ttk, messagebox
from dataclasses import dataclass
from typing import Optional, List, Tuple

import openpyxl

EXCEL_FILE = 'adatok.xlsx'


@dataclass
class Record:
    topic: str
    year_raw: str
    event: str
    mandatory: bool
    start_year: int
    end_year: int


DASHES = ['–', '—', '−', '‑']


def normalize_text(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    for d in DASHES:
        text = text.replace(d, '-')
    return text


# Ebből a függvényből a program a szűréshez számpárt készít.
def parse_year_interval(value) -> Tuple[int, int]:
    if value is None or str(value).strip() == '':
        raise ValueError('Hiányzó évszám')

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        y = int(value)
        return y, y

    text = normalize_text(value)

    # Egyetlen évszám
    if re.fullmatch(r'-?\d+', text):
        y = int(text)
        return y, y

    # Tartomány: pl. 1914-1918, -44--43, 1848-49
    m = re.fullmatch(r'(-?\d+)\s*-\s*(-?\d+)', text)
    if m:
        left_raw, right_raw = m.group(1), m.group(2)
        left = int(left_raw)
        right = int(right_raw)

        # Rövidített második évszám kezelése: 1848-49 -> 1849
        if not right_raw.startswith('-') and len(right_raw) < len(left_raw.lstrip('-')) and left >= 0:
            prefix_len = len(left_raw) - len(right_raw)
            prefix = left_raw[:prefix_len]
            right = int(prefix + right_raw)

        if left <= right:
            return left, right
        return right, left

    raise ValueError(f'Nem értelmezhető évszám: {value}')


# Megjelenítéshez szebb formátum
# Negatív érték esetén: 44 Kr.e.
def display_year(year_raw: str) -> str:
    text = normalize_text(year_raw)
    m = re.fullmatch(r'(-?\d+)\s*-\s*(-?\d+)', text)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return f'{format_single_year(a)}–{format_single_year(b)}'
    if re.fullmatch(r'-?\d+', text):
        return format_single_year(int(text))
    return text



def format_single_year(y: int) -> str:
    return f'{abs(y)} Kr.e.' if y < 0 else str(y)



def load_records(filename: str) -> List[Record]:
    wb = openpyxl.load_workbook(filename)
    ws = wb[wb.sheetnames[0]]
    records: List[Record] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        topic = normalize_text(row[0]).lower()
        year_raw = normalize_text(row[1])
        event = normalize_text(row[2])
        mandatory_cell = normalize_text(row[3]).lower()

        if not topic and not year_raw and not event:
            continue
        if not topic or not year_raw or not event:
            raise ValueError(f'Hiányos sor a(z) {idx}. sorban.')
        if topic not in ('magyar', 'egyetemes'):
            raise ValueError(f'Ismeretlen témakör a(z) {idx}. sorban: {topic}')

        start_year, end_year = parse_year_interval(year_raw)
        records.append(
            Record(
                topic=topic,
                year_raw=year_raw,
                event=event,
                mandatory=(mandatory_cell == 'x'),
                start_year=start_year,
                end_year=end_year,
            )
        )

    if not records:
        raise ValueError('Az Excel fájl nem tartalmaz feldolgozható adatot.')
    return records



def overlaps(record: Record, from_year: Optional[int], to_year: Optional[int]) -> bool:
    if from_year is None and to_year is None:
        return True
    lo = from_year if from_year is not None else -10**9
    hi = to_year if to_year is not None else 10**9
    return record.end_year >= lo and record.start_year <= hi


class QuizApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title('Történelem évszám- és eseménygyakorló')
        self.root.geometry('920x620')

        try:
            self.all_records = load_records(EXCEL_FILE)
        except Exception as e:
            messagebox.showerror('Hiba', f'Nem sikerült megnyitni az {EXCEL_FILE} fájlt.\n\n{e}')
            self.root.destroy()
            return

        self.filtered_records: List[Record] = []
        self.current_record: Optional[Record] = None
        self.current_mode: Optional[int] = None
        self.correct_answer = None
        self.correct_index = None
        self.answer_var = tk.StringVar()
        self.feedback_var = tk.StringVar()

        self._build_ui()
        self.apply_filters(show_message=False)
        self.next_question()

    def _build_ui(self):
        top = ttk.LabelFrame(self.root, text='Szűrők és beállítások', padding=12)
        top.pack(fill='x', padx=12, pady=12)

        ttk.Label(top, text='Témakör:').grid(row=0, column=0, sticky='w')
        self.topic_var = tk.StringVar(value='mindkettő')
        ttk.Combobox(top, textvariable=self.topic_var, state='readonly', width=16,
                     values=['mindkettő', 'magyar', 'egyetemes']).grid(row=0, column=1, sticky='w', padx=(6, 18))

        ttk.Label(top, text='Időszak tól:').grid(row=0, column=2, sticky='w')
        self.from_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.from_var, width=12).grid(row=0, column=3, sticky='w', padx=(6, 18))

        ttk.Label(top, text='Időszak ig:').grid(row=0, column=4, sticky='w')
        self.to_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.to_var, width=12).grid(row=0, column=5, sticky='w', padx=(6, 18))

        self.mandatory_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(top, text='Csak kötelező kerettanterv', variable=self.mandatory_var).grid(row=0, column=6, sticky='w')

        ttk.Label(top, text='Feladattípus:').grid(row=1, column=0, sticky='w', pady=(10, 0))
        self.mode_var = tk.StringVar(value='4. Random')
        ttk.Combobox(
            top,
            textvariable=self.mode_var,
            state='readonly',
            width=28,
            values=[
                '1. Jelöld meg a helyes évszámot',
                '2. Írd be a helyes évszámot',
                '3. Jelöld meg a helyes eseményt',
                '4. Random',
            ],
        ).grid(row=1, column=1, columnspan=3, sticky='w', padx=(6, 18), pady=(10, 0))

        ttk.Button(top, text='Szűrés alkalmazása', command=self.on_filter_click).grid(row=1, column=4, padx=6, pady=(10, 0))
        ttk.Button(top, text='Új kérdés', command=self.next_question).grid(row=1, column=5, padx=6, pady=(10, 0))
        ttk.Button(top, text='Válasz ellenőrzése', command=self.check_answer).grid(row=1, column=6, padx=6, pady=(10, 0))

        info_frame = ttk.Frame(self.root, padding=(12, 0))
        info_frame.pack(fill='x')
        self.pool_label = ttk.Label(info_frame, text='Találatok: 0')
        self.pool_label.pack(anchor='w')

        main = ttk.LabelFrame(self.root, text='Feladat', padding=16)
        main.pack(fill='both', expand=True, padx=12, pady=12)

        self.question_label = ttk.Label(main, text='', font=('Arial', 15, 'bold'), wraplength=820, justify='left')
        self.question_label.pack(anchor='w', pady=(0, 18))

        self.sub_label = ttk.Label(main, text='', font=('Arial', 12), wraplength=820, justify='left')
        self.sub_label.pack(anchor='w', pady=(0, 14))

        self.answer_container = ttk.Frame(main)
        self.answer_container.pack(fill='x', anchor='w')

        self.feedback_label = ttk.Label(main, textvariable=self.feedback_var, font=('Arial', 12, 'bold'), wraplength=820, justify='left')
        self.feedback_label.pack(anchor='w', pady=(18, 8))

        help_text = (
            'Tipp: Kr.e. évszámoknál negatív számot adj meg a szűrőben és a beírásnál is. '
            'Példa: 44 Kr.e. → -44. A program kezeli az évintervallumokat is, például: 1914-1918 vagy 1848-49.'
        )
        ttk.Label(main, text=help_text, wraplength=820, justify='left').pack(anchor='w', pady=(10, 0))

    def parse_optional_year(self, text: str) -> Optional[int]:
        text = normalize_text(text)
        if text == '':
            return None
        if not re.fullmatch(r'-?\d+', text):
            raise ValueError('Az időszűrő mezőkbe csak egyetlen évszám írható, pl. 1526 vagy -44.')
        return int(text)

    def on_filter_click(self):
        self.apply_filters(show_message=True)
        self.next_question()

    def apply_filters(self, show_message: bool = False):
        try:
            from_year = self.parse_optional_year(self.from_var.get())
            to_year = self.parse_optional_year(self.to_var.get())
        except ValueError as e:
            messagebox.showwarning('Érvénytelen szűrő', str(e))
            return

        if from_year is not None and to_year is not None and from_year > to_year:
            messagebox.showwarning('Érvénytelen időintervallum', 'A tól érték nem lehet nagyobb, mint az ig érték.')
            return

        topic = self.topic_var.get()
        only_mandatory = self.mandatory_var.get()

        result = []
        for rec in self.all_records:
            if topic != 'mindkettő' and rec.topic != topic:
                continue
            if only_mandatory and not rec.mandatory:
                continue
            if not overlaps(rec, from_year, to_year):
                continue
            result.append(rec)

        self.filtered_records = result
        self.pool_label.config(text=f'Találatok: {len(self.filtered_records)} rekord')

        if show_message:
            if self.filtered_records:
                messagebox.showinfo('Szűrés kész', f'{len(self.filtered_records)} rekord felel meg a feltételeknek.')
            else:
                messagebox.showwarning('Nincs találat', 'A megadott feltételek mellett nincs megfelelő rekord.')

    def clear_answer_widgets(self):
        for widget in self.answer_container.winfo_children():
            widget.destroy()

    def determine_mode(self) -> int:
        text = self.mode_var.get()
        if text.startswith('1.'):
            return 1
        if text.startswith('2.'):
            return 2
        if text.startswith('3.'):
            return 3
        return random.choice([1, 2, 3])

    def next_question(self):
        self.feedback_var.set('')
        self.answer_var.set('')
        self.clear_answer_widgets()

        if not self.filtered_records:
            self.question_label.config(text='Nincs megjeleníthető kérdés.')
            self.sub_label.config(text='Állíts be más szűrőket, majd kattints a „Szűrés alkalmazása” gombra.')
            self.current_record = None
            self.current_mode = None
            return

        self.current_record = random.choice(self.filtered_records)
        self.current_mode = self.determine_mode()

        if self.current_mode == 1:
            self.build_multiple_choice_year()
        elif self.current_mode == 2:
            self.build_typed_year()
        else:
            self.build_multiple_choice_event()

    def build_multiple_choice_year(self):
        rec = self.current_record
        self.question_label.config(text='Jelöld meg a helyes évszámot!')
        self.sub_label.config(text=f'Esemény: {rec.event}')

        correct = rec.year_raw
        options = {correct}
        pool = [r.year_raw for r in self.filtered_records if r.year_raw != correct]
        random.shuffle(pool)
        for item in pool:
            options.add(item)
            if len(options) == 4:
                break

        options = list(options)
        random.shuffle(options)
        self.correct_answer = correct

        self.choice_var = tk.IntVar(value=-1)
        for i, opt in enumerate(options):
            ttk.Radiobutton(
                self.answer_container,
                text=display_year(opt),
                variable=self.choice_var,
                value=i,
            ).pack(anchor='w', pady=4)
        self.correct_index = options.index(correct)

    def build_typed_year(self):
        rec = self.current_record
        self.question_label.config(text='Írd be a helyes évszámot!')
        self.sub_label.config(text=f'Esemény: {rec.event}')
        self.correct_answer = rec.year_raw

        ttk.Label(self.answer_container, text='Válasz:').pack(anchor='w')
        entry = ttk.Entry(self.answer_container, textvariable=self.answer_var, width=30)
        entry.pack(anchor='w', pady=6)
        entry.focus_set()
        ttk.Label(
            self.answer_container,
            text='Intervallum esetén így írd be: 1914-1918 vagy 1848-49. Kr.e. példa: -44',
        ).pack(anchor='w')

    def build_multiple_choice_event(self):
        rec = self.current_record
        self.question_label.config(text='Jelöld meg a helyes eseményt!')
        self.sub_label.config(text=f'Évszám: {display_year(rec.year_raw)}')

        correct = rec.event
        options = {correct}
        pool = [r.event for r in self.filtered_records if r.event != correct]
        random.shuffle(pool)
        for item in pool:
            options.add(item)
            if len(options) == 4:
                break

        options = list(options)
        random.shuffle(options)
        self.correct_answer = correct

        self.choice_var = tk.IntVar(value=-1)
        for i, opt in enumerate(options):
            ttk.Radiobutton(
                self.answer_container,
                text=opt,
                variable=self.choice_var,
                value=i,
            ).pack(anchor='w', pady=4)
        self.correct_index = options.index(correct)

    def answers_equal(self, user_input: str, correct: str) -> bool:
        try:
            ui = parse_year_interval(user_input)
            ca = parse_year_interval(correct)
            return ui == ca
        except Exception:
            return normalize_text(user_input).lower() == normalize_text(correct).lower()

    def check_answer(self):
        if not self.current_record or not self.current_mode:
            return

        if self.current_mode in (1, 3):
            selected = self.choice_var.get()
            if selected == -1:
                messagebox.showinfo('Hiányzó válasz', 'Előbb jelölj meg egy választ!')
                return
            ok = selected == self.correct_index
        else:
            user_input = self.answer_var.get().strip()
            if not user_input:
                messagebox.showinfo('Hiányzó válasz', 'Írd be a válaszodat!')
                return
            ok = self.answers_equal(user_input, self.correct_answer)

        if ok:
            self.feedback_var.set('Helyes válasz!')
        else:
            right_year = display_year(self.current_record.year_raw)
            right_event = self.current_record.event
            self.feedback_var.set(f'Helytelen. Helyes megoldás: {right_year} – {right_event}')


def main():
    root = tk.Tk()
    try:
        style = ttk.Style()
        if 'vista' in style.theme_names():
            style.theme_use('vista')
    except Exception:
        pass
    QuizApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
